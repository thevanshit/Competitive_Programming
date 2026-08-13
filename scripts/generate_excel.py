#!/usr/bin/env python3
"""
generate_excel.py — Export the ICPC+GSoC 2027 master plan as a formatted Excel workbook.

Builds roadmap/icpc_gsoc_2027_roadmap.xlsx from the data in
roadmap/icpc_gsoc_2027_roadmap.md. Sheets: Overview, Fixed Dates, Phase A-E,
GSoC Books, CP Books, Weekly Template, Contest Strategy, Progress Trackers.

Usage:
    python3 scripts/generate_excel.py
       (Run from the repo root, or script auto-detects it.)
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_FILE = REPO_ROOT / "roadmap" / "icpc_gsoc_2027_roadmap.xlsx"

# ── Palette ────────────────────────────────────────────────────────
NAVY = "1F3864"
BLUE = "2E75B6"
LIGHT = "DEEBF7"
GOLD = "FFF2CC"
RED = "F8CBAD"
GREEN = "E2EFDA"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet(ws, headers, widths, title):
    ws.freeze_panes = "A3"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[2].height = 20
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def section_title(ws, row, text, span, color=NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(vertical="center")
    return row + 1


def put(ws, r, c, val, bold=False, wrap=True, fill=None, merge=None):
    if merge:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge)
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def table(ws, start_row, rows, bold_first=True, alt=True, highlight=()):
    for i, row in enumerate(rows):
        r = start_row + i
        fill = GOLD if (i + 1) in highlight else (LIGHT if alt and i % 2 == 1 else None)
        for j, val in enumerate(row, start=1):
            put(ws, r, j, val, bold=(bold_first and j == 1), fill=fill)
    return start_row + len(rows)


# ── Data ───────────────────────────────────────────────────────────
FIXED_DATES = [
    ("Aug 15, 2026", "START: ICPC + GSoC 2027 Master Plan", "Both", ""),
    ("Aug 20, 2026", "Recruit 2 teammates", "ICPC", "Start today - same university"),
    ("Aug 27, 2026", "Confirm faculty coach", "ICPC", "Ask CS/HOD department"),
    ("Sep 15, 2026", "Register team + submit payment", "ICPC", "Before Sep 21 cutoff"),
    ("Sep 21, 2026", "DEADLINE: ICPC India registration closes", "ICPC", "Team + coach + payment"),
    ("Sep 25, 2026", "CP CURRICULUM COMPLETE", "CP", "Books, CP-31 100%, CSES core, ~20 contests"),
    ("~Oct 2026", "ICPC 2026-27 India Preliminary (online)", "ICPC", "Peak form + team practice"),
    ("Dec 11-12, 2026", "ICPC Chennai Regional", "ICPC", ""),
    ("Dec 18-19, 2026", "ICPC Dhaka Regional", "ICPC", ""),
    ("Dec 22-23, 2026", "ICPC Kanpur Regional", "ICPC", ""),
    ("Dec 27-28, 2026", "ICPC Mathura Regional", "ICPC", ""),
    ("Dec 30, 2026", "ICPC Amritapuri Regional", "ICPC", ""),
    ("Jan 1, 2027", "Phase C: GSoC application season begins", "GSoC", "Stage 3 (ML & DL) + org research"),
    ("Jan 31, 2027", "Org shortlist finalized", "GSoC", "8-10 orgs -> 3 targets"),
    ("Feb 19, 2027", "GSoC 2027 accepted orgs announced", "GSoC", "Start contributions TODAY"),
    ("Mar 10, 2027", "Proposal drafts done", "GSoC", "Feedback from mentors/community"),
    ("Mar 16-31, 2027", "GSoC application window", "GSoC", "Submit EARLY (Mar 16-20)"),
    ("Mar 20-21, 2027", "AWC Championship (if qualified)", "ICPC", "Apps still open until Mar 31"),
    ("Mar 31, 2027", "DEADLINE: GSoC proposals due 18:00 UTC", "GSoC", ""),
    ("Apr 30, 2027", "GSoC 2027 results announced", "GSoC", "Plan for both outcomes"),
    ("May 1-24, 2027", "Community bonding", "GSoC", "Repo dive, env setup, mentor calls"),
    ("May 25, 2027", "GSoC CODING BEGINS", "GSoC", "May 25 - Aug 24"),
    ("Jul 6-10, 2027", "GSoC midterm evaluations", "GSoC", "Milestone 1 delivered"),
    ("Aug 24-31, 2027", "GSoC final evaluations", "GSoC", ""),
    ("Aug 31, 2027", "GSoC 2027 DONE", "GSoC", "Final eval passed"),
    ("Sep 1, 2027", "Phase E: ICPC 2027-28 peak begins", "ICPC", "Team virtuals + old Asia West sets"),
    ("Sep 10, 2027", "Confirm 2027-28 team registration", "ICPC", "Check portal for window"),
    ("Sep 25, 2027", "MISSION COMPLETE", "Both", "Ready for ICPC 2027-28 prelims"),
]

PHASE_A = [
    ("W1", "Aug 15-21", "Foundation repair: sorting, binary search, two pointers, STL",
     "CSES Ch 1-4, CP4 B1 Ch 1-2, CP-Algorithms", "Intro + Sorting & Searching"),
    ("W2", "Aug 22-28", "Dynamic Programming (the #1 gap)",
     "CSES Ch 7, CP4 B1 Ch 3.4 (pp 153-215)", "DP (19 problems)"),
    ("W3", "Aug 29 - Sep 4", "Graphs & Trees: BFS/DFS/Dijkstra/MST/DSU",
     "CSES Ch 11-14, CP4 B1 Ch 4", "Graph (36)"),
    ("W4", "Sep 5-11", "Math: modular arithmetic, combinatorics, number theory",
     "CP4 B1 Ch 5.3-5.6, CSES Ch 21-23", "Mathematics (31)"),
    ("W5", "Sep 12-18", "Advanced DS + Strings: seg tree, BIT, KMP, hashing",
     "CP4 B1 Ch 6, CP4 B2 Ch 9.3-9.8 skim, CSES Ch 9, 26-27", "Range Queries + Strings"),
    ("W6", "Sep 19-25", "PEAK: 1800-1900+ tiers, finals push, review",
     "CP4 B2 Ch 8 skim, notes re-read", "Finish leftovers + re-solves"),
]

PHASE_A_RHYTHM = [
    ("Mon", "Book chapter", "CP-31 + CSES", "Topic practice"),
    ("Tue", "Book chapter", "CP-31 + CSES", "CF contest"),
    ("Wed", "Book chapter", "CP-31 + CSES", "CodeChef Starters"),
    ("Thu", "Book chapter", "CP-31 + CSES", "Upsolve Tue/Wed"),
    ("Fri", "Book chapter", "CP-31 + CSES", "Topic deep-dive"),
    ("Sat", "Reading catch-up", "CP-31 + CSES", "AtCoder ABC"),
    ("Sun", "Reading catch-up", "CP-31 finish tier", "Virtual CF Div2 + review"),
]

PHASE_A_PEAK = [
    ("Sep 19 (Sat)", "CP-31 1800 completion + ABC contest"),
    ("Sep 20 (Sun)", "Virtual CF Div2 + full upsolve"),
    ("Sep 21 (Mon)", "CF contest + upsolve (ICPC registration deadline - done already!)"),
    ("Sep 22 (Tue)", "CP-31 1900 tier + CSES leftovers"),
    ("Sep 23 (Wed)", "CodeChef Starters + upsolve"),
    ("Sep 24 (Thu)", "Final review - notes, re-solve 10 failed problems, template audit"),
    ("Sep 25 (Fri)", "FINAL DAY - CP curriculum 100% complete. Peak form for ICPC prelim."),
]

PHASE_A_RATINGS = [
    ("Aug 23", "1400+ CF"),
    ("Sep 6", "1500+ CF (Specialist)"),
    ("Sep 19", "1600+ CF"),
    ("Sep 25", "1650-1750 CF, AtCoder ~1000-1200, CodeChef ~1850-1950"),
]

PHASE_A_WORKLOAD = [
    ("CP books (CP4 B1 + B2 selected + CSES Handbook)", "977 pages", "~23 pages/day"),
    ("CP-31 remaining (~170 problems)", "170", "~4/day"),
    ("CSES core", "~150", "~3-4/day"),
    ("Contests (5/week x 6 weeks)", "~20", "-"),
    ("Upsolving", "~60", "~1-2/day"),
    ("Total problems", "~400-500", "8-10/day"),
]

PHASE_B_ICPC = [
    ("Prelim (~Oct 2026)",
     "Full team practice the week before - 3 virtual contests + strategy review"),
    ("Post-prelim",
     "If qualified: 6-8 weeks team training for onsite (Dec). 2 team virtuals/week, "
     "role specialization (implementation / math-geometry / graphs-DP), team templates in templates/"),
    ("Onsite (Dec 11-28)", "Perform. Goal: top 20% in your site."),
    ("CP maintenance",
     "2-3h/day, 3 rated contests/week, upsolve only the hardest miss, polish 2 weakest tags"),
]

PHASE_B_GSOC = [
    ("Head First Git", "Do the exercises - git fluency is non-negotiable", "Oct 15"),
    ("Python Cookbook / Fluent Python", "Python 3 proficiency (pick Fluent for depth)", "Nov 15"),
    ("Python Testing with pytest", "Test culture - orgs expect it", "Nov 30"),
    ("Program Management for Open Source", "Understand OSS contribution norms", "Dec 15"),
    ("Practical Statistics for Data Scientists", "Stats foundations", "Dec 31"),
    ("Causal Inference in Python", "Causality (skim chapters 1-3)", "Dec 31"),
]

PHASE_C_JAN = [
    ("Jan 1-7", "Stage 3: Hands-On ML (Geron) ch 1-6", "2 contests + upsolve"),
    ("Jan 8-14", "Stage 3: Deep Learning from Scratch / PyTorch book", "2 contests + upsolve"),
    ("Jan 15-21", "Stage 3: finish + start org research (shortlist 8-10 ML orgs)", "2 contests"),
    ("Jan 22-31", "Choose 3 target orgs (Hugging Face, Jupyter, PyTorch, TensorFlow, ML4SCI, Red Hen Lab)", "2 contests"),
]

PHASE_C_FEB = [
    ("Feb 19", "Accepted orgs announced -> join Discord/Matrix, read 2027 idea list"),
    ("Feb 19 - Mar 15", "CONTRIBUTE: 2-5 PRs per chosen org (bug fixes, docs, good-first-issues, review others' PRs)"),
    ("Feb - Mar", "Stage 4 books in parallel: NLP with Transformers, Hands-On LLMs, Build an LLM from Scratch"),
]

PHASE_C_MAR = [
    ("Mar 1-7", "Write 2-4 proposals: problem -> approach -> timeline -> deliverables"),
    ("Mar 8-15", "Get feedback on drafts from org mentors/community (engage publicly)"),
    ("Mar 16-31", "Submit on GSoC portal. Submit EARLY (Mar 16-20), not on deadline day"),
    ("Mar 20-21", "AWC Championship (weekend - if qualified) - apps stay open until Mar 31"),
]

PHASE_D = [
    ("May 1-24 (bonding)", "Deep repo dive, dev env, milestones with mentor, first small PR merged"),
    ("May 25 - Jul 10", "Milestone 1 (40% of project) - code + tests + docs, weekly mentor syncs"),
    ("Jul 6-10", "Midterm eval - pass, then Milestone 2"),
    ("Jul 11 - Aug 24", "Milestone 2-3 - main feature complete, merge, polish"),
    ("Aug 24-31", "Final eval - demo, docs, final merge -> GSoC 2027 DONE"),
]

PHASE_E = [
    ("Sep 1-7", "Rest/recovery week (post-GSoC), light virtuals, update templates"),
    ("Sep 8-14", "Team mode: 3 ICPC-format virtuals + role drills + speed optimization"),
    ("Sep 15-21", "Mixed-topic marathons: old ICPC Asia West sets (Amritapuri/Kanpur archives)"),
    ("Sep 22-25", "Final review: notes, 10 re-solves, sleep discipline, ready for prelims"),
]

GSOC_BOOKS = [
    ("01 Git, Python & OSS", "Head First Git, Fluent Python, Python Cookbook, pytest, Robust Python, Essential Math, OSS Program Mgmt (7)", "Deep 4, Skim 3", "Oct-Dec 2026", "Prereq for everything"),
    ("02 Stats & Causality", "Practical Stats, Causal Inference (2)", "Deep 1, Skim 1", "Dec 2026", "ML foundations"),
    ("03 ML & DL", "Hands-On ML, DL from Scratch, Fund. of DL, DL with PyTorch, Interpretable ML, Forecasting, Recommender (7)", "Deep 3, Skim 2, Ref 2", "Jan-Feb 2027", "Proposal credibility"),
    ("04 Transformers & LLM", "NLP with Transformers, Hands-On LLMs, Generative DL, AI Engineering, Build LLM from Scratch (5)", "Deep 4, Skim 1", "Feb-Mar 2027", "The ML-org sweet spot"),
    ("05 LLM Apps", "Prompt Eng, Vector DBs, RAG for Production, Designing LLM Apps, GenAI Design Patterns, Context Eng (6)", "Deep 3, Skim 3", "Apr 2027", "Project ideas"),
    ("06 Agents, Evals & Ops", "AI Agents, MCP, LLMOps, Model Eval, LLMs in Production, Observability for LLMs, RL (7)", "Deep 3, Skim 4", "May-Jul 2027", "During coding"),
    ("07 Multimodal & Adv", "VLM, ASR low-resource, AI Systems Perf, DL at Scale, AI Security (5)", "Skim 5", "Jul-Aug 2027", "Skim - awareness"),
    ("08 OS & Networking", "OSTEP, Kurose (2)", "Skim 2", "Aug-Sep 2027", "Interview/ICPC depth"),
    ("09 Backend & Data", "FastAPI, Designing APIs, gRPC, React, SQL, NoSQL, DDIA, Kafka, Spark + more (15)", "Deep 2, Skim 6, Ref 7", "Sep-Nov 2027", "Engineering breadth"),
    ("10 MLOps & Cloud", "Designing ML Systems, Practical MLOps, Docker, K8s, Terraform, GitHub Actions, AWS (10)", "Skim 6, Ref 4", "Nov-Dec 2027", "If not accepted / career"),
    ("11 System Design", "Fundamentals, Hard Parts, Design Patterns, Microservices, System Design Interview 1-2, ML Interviews (10)", "Skim 5, Ref 5", "Jan 2028+", "Interviews - flexible"),
]

CP_BOOKS = [
    ("CSES Handbook", 296, "W1-W4 (Aug 15 - Sep 11)", "~10 pages/day"),
    ("CP4 Book 1", 329, "W1-W5 (Aug 15 - Sep 18)", "~10 pages/day"),
    ("CP4 Book 2 (selected: Ch 8-9 skim)", "~150", "W4-W6 (Sep 5 - Sep 24)", "~6 pages/day"),
]

WEEKDAY_TEMPLATE = [
    ("Morning", "5:30 - 8:00", "2.5h", "Reading (book pages or CP-Algorithms) + notes"),
    ("Classes", "9:00 - 15:00", "-", "College"),
    ("Evening 1", "16:00 - 18:30", "2.5h", "Problem grinding (CP-31 / CSES)"),
    ("Evening 2", "19:00 - 21:00", "2h", "Contest / upsolve / topic deep-dive"),
    ("Review", "21:00 - 21:30", "0.5h", "Notes, template updates"),
]

WEEKEND_TEMPLATE = [
    ("Morning 1", "8:00 - 10:30", "2.5h", "Reading catch-up (missed pages)"),
    ("Morning 2", "10:30 - 13:00", "2.5h", "CP-31 grind (highest remaining tier)"),
    ("Afternoon", "13:00 - 16:00", "3h", "CSES section + long thinking problems"),
    ("Evening", "16:30 - 21:30", "5h", "Contest (Sat: ABC; Sun: Virtual CF Div2) + full upsolve"),
    ("Night", "21:30 - 22:00", "0.5h", "Weekly review + progress tracker update"),
]

HOUR_BUDGET = [
    ("Mon-Fri morning reading (5 x 2.5)", 12.5),
    ("Mon-Fri evening grinding (5 x 2.5)", 12.5),
    ("Mon-Fri contests/upsolve (3 x 2)", 6),
    ("Weekend days (2 x 8-10)", "16-20"),
    ("Total", "~43-47h"),
]

PLATFORMS = [
    ("Codeforces", "Every Div2/Div3/Div4 (Tue/Thu most weeks)", "Solve A-C in 75 min, attack D"),
    ("CodeChef", "Wed Starters", "4-5 problems, top-half finish"),
    ("AtCoder", "Sat ABC", "A-D by Dec 2026, +E by mid-2027"),
    ("ICPC virtuals", "1-2/week in team season", "ICPC-format reps"),
]

CP_TRACKER = [
    ("CP-31 (341)", "~50%", "70%", "90%", "100%"),
    ("CSES (core ~150)", "~5", "60", "110", "~150"),
    ("Books pages", "0", "450", "750", "977"),
    ("Contests", "0", "10", "15", "~20"),
    ("CF rating", "1328", "-", "-", "1650-1750"),
]

GSOC_TRACKER = [
    ("3 OSS PRs (any project)", "Dec 31, 2026"),
    ("Org shortlist (8-10 -> 3 targets)", "Jan 31, 2027"),
    ("Contribution PRs to target orgs", "Mar 15, 2027"),
    ("2-4 proposals submitted", "Mar 31, 2027"),
    ("Accepted (or fallback engaged)", "Apr 30, 2027"),
    ("Midterm eval passed", "Jul 10, 2027"),
    ("GSoC 2027 done (final eval)", "Aug 31, 2027"),
]

ICPC_TRACKER = [
    ("Team registered + paid", "Sep 21, 2026"),
    ("2026-27 prelim: qualified", "Oct 2026"),
    ("2026-27 onsite regional: top 20% site", "Dec 2026"),
    ("AWC Championship (if qualified)", "Mar 20-21, 2027"),
    ("2027-28 prelim: peak form", "Sep-Oct 2027"),
]

WEEKLY_CHECKIN = [
    "Books: ____ pages this week (min 100)",
    "Problems: ____ solved (min 35)",
    "Contests: ____ / planned ____, Upsolved: ____",
    "Notes + templates updated?",
    "GSoC track: ____ progress toward current milestone",
    "Rating: CF ____, AtCoder ____, CodeChef ____",
    "Next week's 3 focus items written down",
]

RULES = [
    "Sep 25, 2026 is sacred - CP curriculum complete. Nothing moves it.",
    "Contests > practice > reading when time is short (all phases).",
    "45-minute rule: struggle 30-45 min, then editorial, then code it yourself. Never skip.",
    "Never skip upsolving. A contest without upsolve is 50% wasted.",
    "GSoC contributions > GSoC books after Feb 19, 2027. PRs win selections.",
    "Team > solo from Sep 2026 - ICPC is 3-persons-1-machine; practice it.",
    "No videos. Text resources only (existing rule).",
    "Sleep 7+ hours. A tired brain loses every contest.",
    "Sunday review is non-negotiable - 10 minutes, keeps all 3 tracks honest.",
    "If you fall behind 3+ days: cut GSoC skim books first, never contests, never notes.",
]


# ── Build workbook ─────────────────────────────────────────────────
def build() -> None:
    wb = Workbook()

    # 1. Overview
    ws = wb.active
    ws.title = "Overview"
    style_sheet(ws, ["Item", "Value"], [34, 110],
                "ICPC + GSoC 2027 MASTER PLAN - Aug 15, 2026 to Sep 25, 2027")
    r = section_title(ws, 3, "MISSION", 2)
    put(ws, r, 1, "Finish CP curriculum by Sep 25, 2026, compete in both ICPC seasons "
        "(2026-27 and 2027-28), and complete GSoC 2027 (ML/LLM track) - while attending "
        "college (classes 9am-3pm Mon-Fri).", merge=2)
    ws.row_dimensions[r].height = 40
    r += 2
    r = section_title(ws, r, "TIME MATH (40-45h/week)", 2)
    for label, val in [
        ("Weekday", "~5h (morning 2.5h + evening 2.5h + contest nights extra)"),
        ("Weekend", "~8-10h/day"),
        ("Weekly total", "~43h = lighter-schedule tier of CP roadmap"),
        ("Daily minimum", "15 pages + 5 problems + 30 min notes"),
    ]:
        put(ws, r, 1, label, bold=True)
        put(ws, r, 2, val)
        r += 1
    r += 1
    r = section_title(ws, r, "REALITY CHECK", 2, color=RED)
    for k, v in [
        ("CP curriculum by Sep 25, 2026", "Doable at 40-45h/week (lighter-schedule variant)"),
        ("ICPC 2026-27 prelim (~Oct 2026)", "6 weeks of sprint feeds directly in"),
        ("ICPC 2026-27 regional (Dec 2026)", "First real attempt - aim for medal-worthy"),
        ("GSoC 2027 application (Mar 31, 2027)", "Requires Stages 1-4 books + PRs by Feb-Mar"),
        ("GSoC 2027 coding (May 25 - Aug 24)", "Acceptance ~10-15% - plan for both outcomes"),
        ("ICPC 2027-28 prelim (~Sep-Oct 2027)", "Full year of depth behind you"),
    ]:
        put(ws, r, 1, k, bold=True)
        put(ws, r, 2, v)
        r += 1

    # 2. Fixed Dates
    ws = wb.create_sheet("Fixed Dates")
    style_sheet(ws, ["Date", "Event", "Track", "What you must have done"],
                [20, 55, 10, 45], "ANCHOR CALENDAR - ALL HARD DEADLINES")
    table(ws, 3, FIXED_DATES, highlight=(5, 12, 19))

    # 3. Phase A
    ws = wb.create_sheet("Phase A - CP Sprint")
    style_sheet(ws, ["Week", "Dates", "Focus", "Books / Articles", "CSES Section"],
                [8, 15, 48, 48, 26], "PHASE A - CP SPRINT (Aug 15 to Sep 25, 2026) - 6 WEEKS")
    r = section_title(ws, 3, "WEEK-BY-WEEK PHASES", 5)
    r = table(ws, r, PHASE_A, highlight=(6,))
    r += 1
    r = section_title(ws, r, "WEEKLY RHYTHM", 5, color=BLUE)
    put(ws, r, 1, "Day", bold=True, fill=BLUE)
    put(ws, r, 2, "Morning (read)", bold=True, fill=BLUE)
    put(ws, r, 3, "Day (grind)", bold=True, fill=BLUE)
    put(ws, r, 4, "Evening", bold=True, fill=BLUE)
    r += 1
    r = table(ws, r, PHASE_A_RHYTHM)
    r += 1
    r = section_title(ws, r, "WEEK 6 PEAK DAYS (Sep 19-25)", 5, color=GOLD)
    for day, act in PHASE_A_PEAK:
        put(ws, r, 1, day, bold=True, fill=GOLD)
        put(ws, r, 2, act, merge=5, fill=GOLD)
        r += 1
    r += 1
    r = section_title(ws, r, "RATING CHECKPOINTS (realistic)", 5, color=BLUE)
    for by, ms in PHASE_A_RATINGS:
        put(ws, r, 1, by, bold=True)
        put(ws, r, 2, ms, merge=5)
        r += 1
    r += 1
    r = section_title(ws, r, "WORKLOAD BUDGET (42 days)", 5)
    for i, (w_, tot, day) in enumerate(PHASE_A_WORKLOAD):
        fill = GREEN if i == len(PHASE_A_WORKLOAD) - 1 else None
        put(ws, r, 1, w_, bold=True, fill=fill)
        put(ws, r, 2, tot, fill=fill)
        put(ws, r, 3, day, fill=fill)
        r += 1

    # 4. Phase B
    ws = wb.create_sheet("Phase B - ICPC Season")
    style_sheet(ws, ["Item", "Detail"], [40, 90],
                "PHASE B - ICPC 2026-27 SEASON + GSoC STAGE 1-2 (Sep 26 to Dec 31, 2026)")
    r = section_title(ws, 3, "ICPC TRACK (priority)", 2)
    r = table(ws, r, PHASE_B_ICPC)
    r += 1
    r = section_title(ws, r, "GSoC STAGE 1-2 (fills weekends + slack)", 2, color=BLUE)
    put(ws, r, 1, "Book", bold=True, fill=BLUE)
    put(ws, r, 2, "Role", bold=True, fill=BLUE)
    put(ws, r, 3, "Finish by", bold=True, fill=BLUE)
    r += 1
    for i, (book, role, fin) in enumerate(PHASE_B_GSOC):
        fill = LIGHT if i % 2 == 1 else None
        put(ws, r, 1, book, bold=True, fill=fill)
        put(ws, r, 2, role, fill=fill)
        put(ws, r, 3, fin, fill=fill)
        r += 1
    r += 1
    r = section_title(ws, r, "MILESTONE", 2, color=GREEN)
    put(ws, r, 1, "By Jan 1, 2027: 3 merged PRs to any OSS project (docs/bug-fixes count) "
        "+ comfortable Python.", merge=2)

    # 5. Phase C
    ws = wb.create_sheet("Phase C - GSoC Apply")
    style_sheet(ws, ["Period", "GSoC work", "CP work"], [16, 80, 25],
                "PHASE C - GSoC APPLICATION SEASON (Jan 1 to Mar 31, 2027)")
    r = section_title(ws, 3, "JANUARY - FOUNDATION FOR PROPOSALS", 3)
    r = table(ws, r, PHASE_C_JAN)
    r += 1
    r = section_title(ws, r, "FEBRUARY - ORG LIST + CONTRIBUTIONS (the differentiator)", 3, color=BLUE)
    r = table(ws, r, PHASE_C_FEB)
    r += 1
    r = section_title(ws, r, "MARCH - PROPOSALS + APPLICATION", 3, color=GOLD)
    r = table(ws, r, PHASE_C_MAR)
    r += 1
    r = section_title(ws, r, "PROPOSAL CHECKLIST", 3, color=GREEN)
    put(ws, r, 1, "Title + problem statement + technical approach + weekly milestone timeline + "
        "expected deliverables + prior contributions (list PRs) + backup plan. 1-2 pages, concrete.",
        merge=3)

    # 6. Phase D
    ws = wb.create_sheet("Phase D - GSoC Coding")
    style_sheet(ws, ["Period", "Focus"], [24, 90],
                "PHASE D - GSoC CODING PERIOD (Apr 1 to Aug 31, 2027)")
    r = section_title(ws, 3, "IF ACCEPTED (the primary plan)", 2)
    r = table(ws, r, PHASE_D)
    r += 1
    r = section_title(ws, r, "IF NOT ACCEPTED (the resilient plan)", 2, color=BLUE)
    for text in [
        "Continue the curriculum at full pace: Stages 5-11 over Apr-Aug (deep-read 5-8, skim 9-11)",
        "Keep 2-3 PRs/month to a chosen org (long-term contributor - huge for next cycle)",
        "Build 2 portfolio projects from the stage books (e.g., RAG app, agent with MCP, eval harness)",
        "CP: 2-3 contests/week + advanced topics (flows, FFT, heavy-light)",
    ]:
        put(ws, r, 1, text, merge=2)
        r += 1
    r += 1
    r = section_title(ws, r, "CP UNDER EITHER OUTCOME", 2, color=GREEN)
    for text in [
        "Accepted: 1 contest/week, upsolve 1 problem. GSoC is the job.",
        "Not accepted: 3 contests/week + topic weeks. Peak by Sep 2027.",
    ]:
        put(ws, r, 1, text, merge=2)
        r += 1

    # 7. Phase E
    ws = wb.create_sheet("Phase E - ICPC Peak")
    style_sheet(ws, ["Week", "Focus"], [16, 90],
                "PHASE E - ICPC 2027-28 PEAK (Sep 1 to Sep 25, 2027)")
    r = section_title(ws, 3, "WEEKLY FOCUS", 2)
    r = table(ws, r, PHASE_E)

    # 8. GSoC Books
    ws = wb.create_sheet("GSoC Books (77)")
    style_sheet(ws, ["Stage", "Books", "Tier", "Window", "Why"],
                [26, 80, 20, 15, 28], "GSoC 77-BOOK STAGE SCHEDULE")
    put(ws, 3, 1, "Legend: Deep = deep-read (primary), Skim = selected chapters, "
        "Ref = reference (look up when needed). Stages 7-11 are deliberately post-GSoC - "
        "they serve interviews/career, not the program.", merge=5)
    ws.row_dimensions[3].height = 28
    r = section_title(ws, 4, "STAGES", 5)
    table(ws, r, GSOC_BOOKS, highlight=(4, 9, 10, 11))

    # 9. CP Books
    ws = wb.create_sheet("CP Books")
    style_sheet(ws, ["Book", "Pages", "When", "Daily share"], [45, 10, 30, 15],
                "CP BOOK SCHEDULE (977 pages -> Phase A)")
    table(ws, 3, CP_BOOKS)

    # 10. Weekly Template
    ws = wb.create_sheet("Weekly Template")
    style_sheet(ws, ["Block", "Time", "Duration", "Activity"], [16, 18, 10, 60],
                "WEEKLY SCHEDULE TEMPLATE (IST)")
    r = section_title(ws, 3, "SCHOOL WEEK (Mon-Fri)", 4)
    r = table(ws, r, WEEKDAY_TEMPLATE)
    r += 1
    r = section_title(ws, r, "WEEKEND (Sat-Sun) - the heavy lift", 4, color=BLUE)
    r = table(ws, r, WEEKEND_TEMPLATE)
    r += 1
    r = section_title(ws, r, "WEEKLY HOUR BUDGET", 4, color=GREEN)
    for i, (src, hrs) in enumerate(HOUR_BUDGET):
        fill = GREEN if i == len(HOUR_BUDGET) - 1 else None
        put(ws, r, 1, src, bold=True, fill=fill)
        put(ws, r, 2, hrs, fill=fill)
        r += 1
    r += 1
    put(ws, r, 1, "Contest slots: Codeforces ~20:05/22:35 IST (Tue/Thu, replaces Evening 2) - "
        "CodeChef Starters Wed 20:00 IST - AtCoder ABC Sat 17:30 IST. Daily minimum: "
        "15 pages + 5 problems + 30 min notes.", merge=4)
    ws.row_dimensions[r].height = 30

    # 11. Contest Strategy
    ws = wb.create_sheet("Contest Strategy")
    style_sheet(ws, ["Platform", "Cadence", "Target"], [20, 45, 45],
                "CONTEST STRATEGY + TEAM FORMATION")
    r = section_title(ws, 3, "PLATFORM CADENCE (all year)", 3)
    r = table(ws, r, PLATFORMS)
    r += 1
    r = section_title(ws, r, "ICPC TEAM - URGENT (do this week)", 3, color=RED)
    for text in [
        "[ ] Recruit 2 teammates from your college (same university = one team)",
        "[ ] Find a faculty coach (ask CS/HOD department - free, just signs up)",
        "[ ] Check your site (Kanpur / Amritapuri / Chennai / Mathura) registration portal",
        "[ ] Register + pay before Sep 21, 2026 (Mathura window: Aug 15 - Sep 21)",
        "[ ] Register for the online prelim (~Oct 2026)",
    ]:
        put(ws, r, 1, text, merge=3)
        r += 1
    r += 1
    r = section_title(ws, r, "GSoC SELECTION ODDS MAXIMIZERS (in order of impact)", 3, color=GREEN)
    for text in [
        "1. Merged PRs before/during application window (Feb 19 - Mar 31) - the #1 predictor",
        "2. Proposal quality - concrete timeline, realistic scope, prior-art cited",
        "3. Community engagement - visible in Discord/Matrix, thoughtful questions on issues",
        "4. Early submission (Mar 16-20) - avoids portal congestion, allows revisions",
    ]:
        put(ws, r, 1, text, merge=3)
        r += 1

    # 12. Progress Trackers
    ws = wb.create_sheet("Progress Trackers")
    style_sheet(ws, ["Metric / Milestone", "Aug 15", "Aug 29", "Sep 12", "Sep 25 goal"],
                [40, 12, 12, 12, 18], "PROGRESS TRACKER")
    r = section_title(ws, 3, "11.1 CP SPRINT (fill daily, review Sundays)", 5)
    r = table(ws, r, CP_TRACKER)
    r += 1
    r = section_title(ws, r, "11.2 GSoC TRACK", 5, color=BLUE)
    put(ws, r, 1, "Milestone", bold=True, fill=BLUE)
    put(ws, r, 2, "Target date", bold=True, fill=BLUE)
    put(ws, r, 5, "Done", bold=True, fill=BLUE)
    r += 1
    for i, (ms, date_) in enumerate(GSOC_TRACKER):
        fill = LIGHT if i % 2 == 1 else None
        put(ws, r, 1, ms, bold=True, fill=fill)
        put(ws, r, 2, date_, fill=fill)
        put(ws, r, 5, "[ ]", fill=fill)
        r += 1
    r += 1
    r = section_title(ws, r, "11.3 ICPC TRACK", 5, color=BLUE)
    put(ws, r, 1, "Milestone", bold=True, fill=BLUE)
    put(ws, r, 2, "Target", bold=True, fill=BLUE)
    put(ws, r, 5, "Done", bold=True, fill=BLUE)
    r += 1
    for i, (ms, date_) in enumerate(ICPC_TRACKER):
        fill = LIGHT if i % 2 == 1 else None
        put(ws, r, 1, ms, bold=True, fill=fill)
        put(ws, r, 2, date_, fill=fill)
        put(ws, r, 5, "[ ]", fill=fill)
        r += 1
    r += 1
    r = section_title(ws, r, "11.4 WEEKLY CHECK-IN (every Sunday, 10 min)", 5, color=GREEN)
    for text in WEEKLY_CHECKIN:
        put(ws, r, 1, "[ ] " + text, merge=5)
        r += 1
    r += 1
    r = section_title(ws, r, "12. RULES OF ENGAGEMENT", 5, color=GOLD)
    for i, rule in enumerate(RULES, start=1):
        put(ws, r, 1, f"{i}. {rule}", merge=5)
        r += 1

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    build()